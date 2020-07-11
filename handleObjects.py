import openpyxl

# file read
dest_filename = 'c:\\tmp\\sample.xlsx' 
wb = openpyxl.load_workbook(dest_filename)
sheet = wb['ksheet']
#print(sheet.cell(row=1, column=1).value)
print("%s" % sheet.max_column)
print("%s" % sheet.max_row)

print('Shapes count : %s ' % len(sheet.Shapes))

row_range = sheet[1:sheet.max_row]
for row in row_range:
  for cell in row:
    print(cell.value)
  print(' ')

# file write
# wb 는 읽은 파일에 그대로 sheet 추가하기
sheet2 = wb.create_sheet('jsheet')
sheet2.title = 'auto write'
#wb.save(filename = dest_filename)
wb.close()
